#These are the library that i need for my code
from tkinter import *
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
from datetime import datetime, timedelta
from tkinter import messagebox
import random

#For my to-do-list main window
root = Tk()
root.title("To-Do-List")
root.geometry("650x650")

#Create an empty list, all of the tasks later will be stored here
tasks = []

#Stores streak and percentage, start from zero
streak = 0
percentage = 0

#Last date the streak updated, since at first there isn't one so = None
last_streak_date = None

#Images
darkmode_bg = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\darkmode.png")
lightmode_bg = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\lightmode.png")

startlight_button = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\startlight.png")
startlight_photo = ImageTk.PhotoImage(startlight_button)

streak_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\streak.png")
streak_img = streak_img.resize((60,60))
streak_photo = ImageTk.PhotoImage(streak_img)

star_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\star.png")
star_img = star_img.resize((70,70))
star_photo = ImageTk.PhotoImage(star_img)

over_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\over.png")
over_img = over_img.resize((65,65))
over_photo = ImageTk.PhotoImage(over_img)

worried_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\scared.png")
worried_img = worried_img.resize((65,65))
worried_photo = ImageTk.PhotoImage(worried_img)

sirine_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\sirine.png")
sirine_img = sirine_img.resize((16, 16))
sirine_photo = ImageTk.PhotoImage(sirine_img)

red_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\redbutton.png")
red_img = red_img.resize((16, 16))
red_photo = ImageTk.PhotoImage(red_img)

yellow_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\yellowbutton.png")
yellow_img = yellow_img.resize((17,17))
yellow_photo = ImageTk.PhotoImage(yellow_img)

green_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\greenbutton.png")
green_img = green_img.resize((16,16))
green_photo = ImageTk.PhotoImage(green_img)

plus_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\plus.png")
plus_img = plus_img.resize((70, 70))
plus_photo = ImageTk.PhotoImage(plus_img)

bello_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\bello.png")
bello_img = bello_img.resize((180, 180))
bello_photo = ImageTk.PhotoImage(bello_img)

happy_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\happy.png")
happy_img = happy_img.resize((180,180))
happy_photo = ImageTk.PhotoImage(happy_img)

char_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\character.png")
char_img = char_img.resize((180,180))
char_photo = ImageTk.PhotoImage(char_img)

sleep_img = Image.open(r"C:\Users\ASUS\.vscode\to-do-list\sleep.png")
sleep_img = sleep_img.resize((180,180))
sleep_photo = ImageTk.PhotoImage(sleep_img)

#Welcome page

#When the user run the code, the welcome page will be in light mode so = False
dark_mode = False

#Create frame that will hold my dark mode button, like a box/container to put my dark mode button 
top_frame = Frame(root)

#The frame stretches in x-axis/horizontally
top_frame.pack(fill = "x")

#Create frame that will hold my welcome page and main page
page_frame = Frame(root)

#The frame stretches in both horizontally & vertically, fill all the available space
page_frame.pack(fill = BOTH, expand = True)

#Create welcome page inside the page_frame
welcome_frame = Frame(
    page_frame, 
    width = 650, 
    height = 650
)
welcome_frame.pack(fill = BOTH, expand = True)

#To make the frame stay 650x650
welcome_frame.pack_propagate(False)

#Theme

#This function use to update my welcome page theme
def update_welcome_theme():
    if dark_mode:
        bg_colour = "#69666c"
        fg_colour = "#FFFFFF"
    else:
        bg_colour = "#f2eebf"
        fg_colour = "#000000"
    
    streakncomp = [
        streak_frame,
        completion_frame,
        streak_text,
        completion_text,
        sphoto_label,
        cphoto_label
    ]

    for all in streakncomp:
        all.config(bg = bg_colour)
    
    titnlab = [
        streak_title,
        completion_title,
        streak_label,
        percentage_label
    ]

    for all in titnlab:
        all.config(bg = bg_colour, fg = fg_colour)

#Create background label inside the welcome_frame and no border around the label
background_label = Label(welcome_frame, borderwidth = 0)

#Place for the background label start from top left corner with 100% width & height of welcome_frame
background_label.place(
    x = 0, 
    y = 0, 
    relwidth = 1, 
    relheight = 1
)

#This function use to resize my welcome page whenever the window changes size
#If there's no event like the window change size, mouse clicked, just use None
def resize_welcome(event = None):
    global background_photo

    if dark_mode:
        img = darkmode_bg
    else:
        img = lightmode_bg

    #Get current width and height of the frame
    width = welcome_frame.winfo_width()
    height = welcome_frame.winfo_height()

    #Resize the background image
    resized = img.resize((width, height))

    #Coverts the image for Tkinter and update the background image
    background_photo = ImageTk.PhotoImage(resized)
    background_label.config(image=background_photo)

    #This put the background label behind other widgets
    background_label.lower()

#When the frame changes size/position, call resize_welcome function
welcome_frame.bind("<Configure>", resize_welcome)

resize_welcome()

#Streak 

if dark_mode:
    bg_colour = "#69666c"
    fg_colour = "#FFFFFF"
else:
    bg_colour = "#f2eebf"
    fg_colour = "#000000"

streak_frame = Frame(
    welcome_frame,
    bg = bg_colour,     
)

#The position and size of the frame. Center of the frame is placed at (0.4, 0.66)
streak_frame.place(
    relx = 0.4,
    rely = 0.66,
    anchor ="center",
    width = 250,
    height = 95
)

sphoto_label = Label(
    streak_frame,
    image = streak_photo,
    bg = bg_colour
)
#Place my streak buddy on the left side of the frame & create space between img and text
sphoto_label.pack(side = LEFT, padx = (0,10))

#Create another frame inside streak_frame for the text and it's next of the buddy
streak_text = Frame(streak_frame, bg = bg_colour)
streak_text.pack(side = LEFT)

#Create the streak title in streak_text
streak_title = Label(
    streak_text,
    text = "Streak:",
    font = ("Helvetica", 14),
    bg = bg_colour,
    fg = fg_colour
)
#Pack the title to the west or the left side so it won't be on the center
streak_title.pack(anchor = "w")

#Create label in streak_text to show current streak
streak_label = Label(
    streak_text, 
    text = f"{streak} Day",
    font = ("Helvetica", 20, "bold"),
    bg = bg_colour,
    fg = fg_colour
)
streak_label.pack(anchor = "w")

#This function use to update the streak shown on the screen. If the streak 1, show 1 Day
def update_streak():
    streak_label.config(
        text = f"{streak} Day" if streak == 1 else f"{streak} Days"
    )

#This function use to check if the streak should increase, stay the same, or restart
def check_streak():
    global streak, last_streak_date

    today = datetime.now().date()

    #Check if there's any task done that day, atleast one task
    completed_today = any(
        task.get("completed_date") == today.strftime("%m-%d-%Y")
        for task in tasks
    )

    #For the streak to won't change unless there's atleast one task done that day
    if not completed_today:
        return

    #If today's streak has already been counted then return/stop
    if last_streak_date == today:
        return

    #Check whether we never had streak before 
    if last_streak_date is None:
        streak = 1
    
    #Check whether to day is one day after our last streak
    elif today == last_streak_date + timedelta(days = 1):
        streak += 1
    
    #If we skip a day/broke our streak, it will restart to 1
    else:
        streak = 1

    #To let the tomorrow program know our previous streak date
    last_streak_date = today

    update_streak()
    save_task()

#Completion Percentage

completion_frame = Frame(
    welcome_frame,
    bg = bg_colour
)

completion_frame.place(
    relx = 0.6,
    rely = 0.66,
    anchor = "center",
    width = 280,
    height = 75,
)

completion_text = Frame(completion_frame, bg=bg_colour)
completion_text.pack(side = LEFT, padx = (0,10))

completion_title = Label(
    completion_text,
    text = "Completion Percentage:",
    font = ("Helvetica",14),
    bg = bg_colour,
    fg = fg_colour
)
#Pack the title to the east or the right side
completion_title.pack(anchor = "e")

percentage_label = Label(
    completion_text,
    text = f"{percentage}%",
    font = ("Helvetica",20,"bold"),
    bg = bg_colour,
    fg = fg_colour
)
percentage_label.pack(anchor="e")

cphoto_label = Label(
    completion_frame,
    image = star_photo,
    bg = bg_colour
)
cphoto_label.pack(side = LEFT, padx = (10,0))

#This function use to calculate our completion percentage/progress
def update_percentage():
    global percentage

    #Count the total tasks and completed tasks
    total = len(tasks)
    completed = sum(task["done"] for task in tasks)

    #If the user hasn't add any tasks, the percentage will be 0
    if total == 0:
        percentage = 0
    
    #if the user has tasks, then count the percentage
    else: 
        percentage = int((completed / total) * 100)

    #Update the welcome page, progress, completed, and remaining
    percentage_label.config(
        text = f"{percentage} %"
    )

    progress_label.config(
        text = f"{percentage}% complete"
    )

    completed_label.config(
        text = f"Completed: {completed}"
    )

    remaining_label.config(
        text = f"Remaining task: {total - completed}"
    )

#Start

#This function use to start when the start button clicked
def start ():

    #Hide the welcome page and show the main page
    welcome_frame.pack_forget()
    main_frame.pack(fill = BOTH, expand = True)

    update_daily_quote()
    update_percentage()
    update_upcoming()
    update_overdue()
    update_buddy()

if dark_mode:
    bg_colour = "#1F1F1F"
else:
    bg_colour = "#FFFFFF"

#Create the start button in welcome_frame and set the position
start_button = Button(
    welcome_frame,
    image = startlight_photo,
    command = start,
    bd = 0,
    highlightthickness=0,
    relief = "flat",
    cursor = "hand2",
    width = 400,
    bg = bg_colour,
    activebackground = bg_colour
)

start_button.place(
    relx = 0.5,
    rely = 0.86,
    anchor = "center"
)

#Dark Mode

#This function use to toggle the theme in main page
def toggle_theme():
    global dark_mode

    dark_mode = not dark_mode

    if dark_mode:
        main_colour = "#1F1F1F"
        second_colour = "#373737"
        text_colour = "#FFFFFF"
        speech_colour = "#3A3A3A"

        root.config(bg = main_colour)

        main = [
            main_frame,
            content_frame, 
            dashboard_frame,
            button_frame,
            top_frame,
            page_frame,
            buddy_frame,
            buddy_label,
            task_container,
            canvas
        ]

        for all in main:
            all.config(bg = main_colour)
        
        second = [
            search_frame,
            task_list_frame,
            overdue_list_frame,
            smallbud_frame,
            overdue_buddy_frame,
            upcoming_container,
            overdue_container,
            upcoming_canvas,
            overdue_canvas
        ]

        for all in second:
            all.config(bg = second_colour)
        
        secntext = [
            time_frame,
            progress_frame,
            upcoming_frame,
            overdue_frame,
            quote_frame,
            progress_label,
            completed_label,
            remaining_label,
            clock_label,
            search_label,
            quote_label,
            search_entry,

            delete_button,
            back_button,
            theme_button,
            search_button,
            show_button
        ]

        for all in secntext:
            all.config(bg = second_colour, fg = text_colour)
        
        speech = [
            speech_label,
            small_speech_label,
            overdue_speech
        ]

        for all in speech:
            all.config(bg = speech_colour, fg = text_colour)
        
        plus_button.config(bg = main_colour, activebackground = main_colour)
        start_button.config(bg = main_colour, activebackground= main_colour)

        smallbud_label.config(bg = second_colour, activebackground = second_colour)
        overdue_buddy.config(bg = second_colour, activebackground= second_colour)
    
    else:
        main_colour = "#FFFFFF"
        second_colour = "#fafdff"
        text_colour = "#000000"
        speech_colour = "#FFF8D6"

        root.config(bg = main_colour)

        main = [
            main_frame,
            content_frame, 
            dashboard_frame,
            button_frame,
            top_frame,
            page_frame,
            buddy_frame,
            buddy_label,
            task_container,
            canvas
        ]

        for all in main:
            all.config(bg = main_colour)
        
        second = [
            search_frame,
            task_list_frame,
            overdue_list_frame,
            smallbud_frame,
            overdue_buddy_frame,
            upcoming_container,
            overdue_container,
            upcoming_canvas,
            overdue_canvas
        ]

        for all in second:
            all.config(bg = second_colour)
        
        secntext = [
            time_frame,
            progress_frame,
            upcoming_frame,
            overdue_frame,
            quote_frame,
            progress_label,
            completed_label,
            remaining_label,
            clock_label,
            search_label,
            quote_label,
            search_entry,
            delete_button,
            back_button,
            theme_button,
            search_button,
            show_button
        ]

        for all in secntext:
            all.config(bg = second_colour, fg = text_colour)
        
        speech = [
            speech_label,
            small_speech_label,
            overdue_speech
        ]

        for all in speech:
            all.config(bg = speech_colour, fg = text_colour)
        
        plus_button.config(bg = main_colour, activebackground = main_colour)
        start_button.config(bg = main_colour, activebackground = main_colour)

        smallbud_label.config(bg = second_colour, activebackground = second_colour)
        overdue_buddy.config(bg = second_colour, activebackground = second_colour)
    
    refresh_tasks()
    resize_welcome()
    update_upcoming()
    update_overdue()
    update_buddy()
    update_welcome_theme()

#Create the theme button in top_frame and the space
theme_button = Button(
    top_frame,
    text = "Dark Mode🌙",
    command = toggle_theme,
    cursor = "hand2"
)
theme_button.pack(pady = 5)

#Main Page

main_frame = Frame(page_frame)

search_frame = Frame(main_frame)
search_frame.pack(fill = X, pady = 10)

content_frame = Frame(main_frame)

#Tells Tkinter how much space each column should take
content_frame.columnconfigure(0, weight = 3)
content_frame.columnconfigure(1, weight = 1)

#Allows it to expand vertically when the window gets taller
content_frame.rowconfigure(0, weight = 1)

#Let the content fill the remaining window & keep the size
content_frame.pack(fill = BOTH, expand = True)
content_frame.grid_propagate(False)

#Left Side

#Create frame for tasks list area, put it in row and column 0 & fills its entire grid cell
list_frame = Frame(content_frame)
list_frame.grid(
    row = 0, 
    column = 0, 
    sticky = "nsew"
)
#Stops the frame from shrinking to fit its widget
list_frame.pack_propagate(False)

#Right Side

dashboard_frame = Frame(content_frame, width = 250)
dashboard_frame.grid(
    row = 0, 
    column = 1, 
    sticky = "nsew"
)
dashboard_frame.grid_propagate(False)

#Progress

progress_frame = LabelFrame(
    dashboard_frame, 
    text = "⚡Progress",
    font = ("Helvetica", 14, "bold"),
    padx = 10,
    pady = 10
)
progress_frame.pack(fill = X, padx = 10, pady = 10)

#Create labels to show progress, completed, & remaining 
progress_label = Label(progress_frame, text = "0% complete")
progress_label.pack()

completed_label = Label(progress_frame, text = "Completed: 0")
completed_label.pack()

remaining_label = Label(progress_frame, text = "Remaining task: 0")
remaining_label.pack()

#Overdue

overdue_frame = LabelFrame(
    dashboard_frame,
    text="⚠️ Overdue Tasks",
    font=("Helvetica", 14, "bold")
)
overdue_frame.pack(fill = X, padx = 10, pady = 10)

#Create overdue canvas so we can scroll on it 
overdue_canvas = Canvas(
    overdue_frame,
    height = 100,
    highlightthickness = 0
)

#Create vertical scroll bar
overdue_scroll = Scrollbar(
    overdue_frame,
    orient = "vertical",
    command = overdue_canvas.yview
)

#Connect the canvas and the scrollbar so they synchronized
overdue_canvas.configure(
    yscrollcommand = overdue_scroll.set
)

#Display them, canvas on the left the scrollbar on the right side of it
overdue_canvas.pack(side = LEFT, fill = BOTH, expand = True)
overdue_scroll.pack(side = RIGHT, fill = Y)

overdue_container = Frame(overdue_canvas)

#Put the frame inside the canvas, starts at top left corner
overdue_canvas.create_window(
    (0, 0),
    window = overdue_container,
    anchor = "nw"
)

#Update scrolling size/region automatically
overdue_container.bind(
    "<Configure>",
    lambda e: overdue_canvas.configure(
        scrollregion = overdue_canvas.bbox("all")
    )
)

#Create frames for the overdue tasks list and my overdue buddy
overdue_list_frame = Frame(overdue_container)
overdue_list_frame.pack(side = LEFT, fill = BOTH, expand = True)

overdue_buddy_frame = Frame(overdue_container)
overdue_buddy_frame.pack(
    side = RIGHT, 
    anchor = "n", 
    padx = (150,10), 
    pady = 5
)

#Create label of speech bubble for my overdue buddy. If the message is longer than 180 pixels, it wraps onto the next line
overdue_speech = Label(
    overdue_buddy_frame,
    text = "No overdue tasks! 🎉",
    font = ("Helvetica",10),
    padx = 15,
    pady = 10,
    wraplength = 180,
    relief = "ridge",
    bd = 2
)
overdue_speech.pack(pady = (0,5))

if dark_mode:
    overdue_speech.config(
        bg = "#3A3A3A",
        fg = "#FFFFFF"
    )
else:
    overdue_speech.config(
        bg = "#FFF8D6",
        fg = "#000000"
    )

overdue_buddy = Label(
    overdue_buddy_frame, 
    image = worried_photo
)
overdue_buddy.pack()

#This function use to find overdue tasks and display it
def update_overdue():
    now = datetime.now()

    overdue = []

    for task in tasks:

        #Ignore completed task and convert date & time into datetime
        if not task["done"]:
            task_datetime = datetime.strptime(
                f"{task['date']} {task['time']}",
                "%m-%d-%Y %H:%M"
            )

            #Check if it overdue and save it to the list
            if task_datetime < now:
                overdue.append((task_datetime, task))

    #Sort overdue tasks by the first value of each tuple, which x[0] means task_datetime
    overdue.sort(key = lambda x: x[0])

    #Clear all widgets
    for widget in overdue_list_frame.winfo_children():
        widget.destroy()

    #Check wheteher the frame is visible or not
    if not overdue_frame.winfo_ismapped():
        overdue_frame.pack(
            before=quote_frame,
            fill = X,
            padx = 10,
            pady = 10
        )

    if dark_mode:
            bg_colour = "#565656"
            fg_colour = "white"
    else:
            bg_colour = "white"
            fg_colour = "black"

    #Display each overdue tasks
    for task_datetime, task in overdue:

        #Calculate how long the tasks overdue
        days = (now.date() - task_datetime.date()).days

        #Status text
        if days == 0:
            status = "Today"
        elif days == 1:
            status = "Yesterday"
        else:
            status = f"{days} days ago"

        #Create small card with sirine icon inside
        sirine = Frame(overdue_list_frame, bg = bg_colour)
        sirine.pack(fill = X, padx = 5, pady = 4)

        sirine_label = Label(
            sirine,
            image = sirine_photo,
            bg = bg_colour
        )

        #Keep image in memory, without this it may disappear
        sirine_label.image = sirine_photo      
        sirine_label.pack(side = LEFT, padx = 5)

        #Show the task information
        Label(
            sirine,
            text = f"{task['task']}\n{status}",
            justify = "left",
            anchor = "w",
            wraplength = 180,
            bg = bg_colour,
            fg = fg_colour,
            width = 18
        ).pack(fill = X, padx = 5, pady = 4)

    #The overdue buddy speech
    if overdue:
        overdue_speech.config(
            text = f"Oh no!\n{len(overdue)} overdue task(s)! 😥"
        )

    else:
        overdue_speech.config(
            text = "Yuhu~\nNo overdue tasks~"
        )

#Quote

quote_frame = LabelFrame(
    dashboard_frame,
    text = "💡 Daily Motivation",
    font = ("Helvetica", 14, "bold"),
    padx = 10,
    pady = 10
)
quote_frame.pack(fill = X, padx = 10, pady = 10)

#Create label where the motivational quote will appear
quote_label = Label(
    quote_frame,
    text = "",
    wraplength = 230,
    justify = "center",
    font = ("Helvetica", 11, "italic")
)
quote_label.pack()

#The list of quotes
quotes = [
    "Small progress is still progress. 🌱",
    "Today's effort is tomorrow's success. 🌸",
    "Believe in yourself . You've got this! 💪",
    "Every checked task is a step forward. 🏃‍♂️",
    "“The secret of getting ahead is getting started.” – Mark Twain",
    "“It does not matter how slowly you go as long as you do not stop.” – Confucius"
]

#This function use to choose today's quote
def update_daily_quote():
    today = datetime.now().date()

    #Set the random seed so it always picks the same quote for a day
    random.seed(today.toordinal())

    #Choose random quote for that day
    quote_label.config(
        text = random.choice(quotes)
    )

#Upcoming Tasks

upcoming_frame = LabelFrame(
    dashboard_frame,
    text = "🌸Upcoming Tasks",
    font = ("Helvetica", 14, "bold")
)
upcoming_frame.pack(fill = X, padx = 10, pady = 10)

upcoming_canvas = Canvas(
    upcoming_frame,
    height = 120,
    highlightthickness = 0
)
upcoming_canvas.pack(side = LEFT, fill = BOTH, expand = True)

upcoming_scroll = Scrollbar(
    upcoming_frame,
    orient = "vertical",
    command = upcoming_canvas.yview
)
upcoming_scroll.pack(side = RIGHT, fill = Y)

upcoming_canvas.configure(yscrollcommand = upcoming_scroll.set)

upcoming_container = Frame(upcoming_canvas)

upcoming_canvas.create_window(
    (0, 0),
    window = upcoming_container,
    anchor = "nw"
)

upcoming_container.bind(
    "<Configure>",
    lambda e: upcoming_canvas.configure(
        scrollregion = upcoming_canvas.bbox("all")
    )
)

task_list_frame = Frame(upcoming_container)
task_list_frame.pack(side = LEFT, fill = BOTH, expand = True)

smallbud_frame = Frame(upcoming_container)
smallbud_frame.pack(side = RIGHT, anchor = "n", padx = (150,10), pady = 5)

small_speech_label = Label(
    smallbud_frame,
    text = "Let's get started! 🌱",
    font = ("Helvetica",10),
    padx = 15,
    pady = 10,
    wraplength = 100,
    relief = "ridge",
    bd = 2
)
small_speech_label.pack(pady = (0,5))

if dark_mode:
    small_speech_label.config(
        bg = "#3A3A3A",
        fg = "#FFFFFF"
    )
else:
    small_speech_label.config(
        bg = "#FFF8D6",
        fg = "#000000"
    )

smallbud_label = Label(
    smallbud_frame,
    image = over_photo
)
smallbud_label.pack()

#This function use to give warning if a task's due is soon
def check_reminders():
    now = datetime.now()

    found = False

    for task in tasks:

        #Ignore/skip completed task
        if task["done"]:
            continue
        
        #Convert due date and time into datetime
        due = datetime.strptime(
            f"{task['date']} {task['time']}",
            "%m-%d-%Y %H:%M"
        )

        #Calculates how many minutes left
        minutes_left = int((due - now).total_seconds() / 60)

        #When the task is due within the time, the buddy will give reminder
        if 0 <= minutes_left <= 30:

            small_speech_label.config(
                text = f"⚠️ Hurry!\n'{task['task']}'\nis due in {minutes_left} minutes!"
            )

            smallbud_label.config(image = over_photo)
            found = True
            break

        elif 30 < minutes_left <= 60:

            small_speech_label.config(
                text = f"⏰ Don't forget!\n'{task['task']}'\nis due in 1 hour!"
            )

            smallbud_label.config(image = over_photo)
            found = True
            break
    
    #If the task due above 1 hour
    if not found:
        small_speech_label.config(
            text = "Your tasks are waiting for you 👻"
        )

    #Wait ten secs amount of time, then call this function again
    root.after(10000, check_reminders)

#This function use to update the upcoming tasks
def update_upcoming():
    now = datetime.now()

    upcoming = []

    for task in tasks:
        if not task["done"]:

            task_datetime = datetime.strptime(
                f"{task['date']} {task['time']}",
                "%m-%d-%Y %H:%M"
            )

            #Add both the datetime and the task to the list
            upcoming.append((task_datetime, task))

    #This assigns numbers to priorities
    priority_order = {
        "High": 0,
        "Medium": 1,
        "Low": 2
    }

    #Sort each task by three rules
    upcoming.sort(
        key = lambda x: (
            
            #Tasks that aren't overdue appear before overdue one
            x[0] < now,

            #Sort by date and time
            x[0],

            #If two tasks have exactly the same due date and time, then priority decide their order
            priority_order[x[1]["priority"]]
        )
    )

    for widget in task_list_frame.winfo_children():
        widget.destroy()

    #If there's no upcoming task
    if not upcoming:

        if dark_mode:
            bg_colour = "#565656"
            fg_colour = "white"
        else:
            bg_colour = "white"
            fg_colour = "black"


        Label(
            task_list_frame,
            text = "Hooray, no upcoming tasks! 🥳",
            bg = bg_colour,
            fg = fg_colour,
        ).pack(pady=10)

        #Stop the function because there are no tasks to display
        return

    for task_datetime, task in upcoming:

        if dark_mode:
            bg_colour = "#565656"
            fg_colour = "white"
        else:
            bg_colour = "white"
            fg_colour = "black"

        #Calculate how many days left
        days = (task_datetime.date() - now.date()).days

        #Display the due date and determine the icon
        if days == 0:
            day = "Today"
        elif days == 1:
            day = "Tomorrow"
        elif days > 1:
            day = f"In {days} days"
        else:
            day = "Overdue"

        if task["priority"] == "High":
            icon = red_photo
        elif task["priority"] == "Medium":
            icon = yellow_photo
        else:
            icon = green_photo

        #Create card for the task
        card = Frame(task_list_frame, bg = bg_colour)
        card.pack(fill = X, padx = 5, pady = 4)

        #Show the icon
        icon_label = Label(
            card,
            image = icon,
            bg = bg_colour
        )
        icon_label.image = icon     
        icon_label.pack(side = LEFT, padx = 5)

        #Display the task information
        Label(
            card,
            text=(
                f"{task['task']}\n"
                f"{day} {task_datetime.strftime('%H:%M')}"
            ),
            justify = "left",
            anchor = "w",
            wraplength = 180,
            bg = bg_colour,
            fg = fg_colour,
            width = 18
        ).pack(fill = X, padx = 5, pady = 4)

#Time

time_frame = LabelFrame(
    dashboard_frame,
    text = "⌛Current Time",
    font = ("Helvetica", 14, "bold"),
    padx = 10,
    pady = 10
)
time_frame.pack(fill = X, padx = 10, pady = 10)

clock_label = Label(
    time_frame,
    font = ("Helvetica", 12)
)
clock_label.pack()

#This function use for live digital clock that update every sec
def update_clock():

    #Get the current time with the format
    current = datetime.now().strftime("%H:%M:%S")

    #Update the clock label on the screen
    clock_label.config(text = current)

    #Update every second, so it keep running continously
    root.after(1000, update_clock)

#Search

search_label = Label(search_frame, 
        text = "Search"
)
search_label.pack(side = LEFT, padx = 10)

search_entry = Entry(search_frame)
search_entry.pack(side = LEFT, fill = X, expand = True, padx = 10)

#When the user press enter, run search_task
search_entry.bind(
    "<Return>",
    lambda event: search_tasks()
)

#Mouse

#This function use to make any canvas scroll with the mouse wheel
def bind_mousewheel(canvas):

    #This function runs when the user scrolls with the mouse wheel
    def scroll_ing (event):

        #Get current scroll position
        first, last = canvas.yview()
        
        #Stop scrolling when the user already at the top
        if event.delta > 0 and first <= 0:
            return
        
        #Scroll the canvas
        canvas.yview_scroll(
            int(-1 * (event.delta / 120)),
            "units"
        )

    #When the mouse enter the canvas, the mouse wheel becomes connected to scroll_ing
    canvas.bind(
        "<Enter>",
        lambda e: canvas.bind_all("<MouseWheel>", scroll_ing)
    )

    #When the mouse leaves the canvas, the mouse wheel is disconnected
    canvas.bind(
        "<Leave>",
        lambda e: canvas.unbind_all("<MouseWheel>")
    )          

#To make upcoming and overdue canvas scrollable with mouse wheel
bind_mousewheel(upcoming_canvas) 
bind_mousewheel(overdue_canvas)   

#Tasks

#task = one task dictionary, number = task number. Like 1. Do math homework
def display_task(task, number):
    if dark_mode:
        text_colour = "white"

        if task["priority"] == "High":
            box_colour = "#5B2C2C"
        elif task["priority"] == "Medium":
            box_colour = "#5C5230"
        else:
            box_colour = "#2E5A38"
    else:
        text_colour = "black"

        if task["priority"] == "High":
            box_colour = "#FFD6D6"
        elif task["priority"] == "Medium":
            box_colour = "#FFF3CD"
        else:
            box_colour = "#D4F8D4"

    #Create the task card and display it
    task_frame = LabelFrame(
        task_container,
        bg = box_colour,
        fg = text_colour,
        relief = "solid",
        bd = 1,
        width = 320,
        height = 140
    )
    task_frame.pack(anchor = "w", padx = (10,20), pady = 5)

    task_frame.pack_propagate(False)
    
    #Save the frame inside the task
    task["frame"] = task_frame

    #Create delete button, when it clicked the delete_task executed
    delete_button = Button(
        task_frame,
        text = "❌",
        command = lambda t=task: delete_task(t),
        cursor = "hand2"
    )

    #Save delete button inside the task
    task["delete_button"] = delete_button

    #Show delete button only when in delete mode
    if delete_mode:
        delete_button.pack(side = RIGHT, padx = 5)

    #Create checkbox variable
    done_var = BooleanVar(value = task["done"])

    #This function run when the checkbox is clicked
    def mark_done(t=task, var=done_var):

        #Update the task status
        t["done"] = var.get()

        if t["done"]:
            
            #If checked, save completion date
            t["completed_date"] = datetime.now().strftime("%m-%d-%Y")

        else:
            #If unchecked, erase the completion date
            t["completed_date"] = ""

        check_streak()

        update_percentage()
        update_upcoming()
        update_overdue()
        update_buddy()

        save_task()

    #Create the checkbox
    Checkbutton(
        task_frame,
        variable = done_var,
        command = mark_done,
        bg = box_colour,
        fg = text_colour,
        activebackground = box_colour,
        selectcolor = box_colour,
        cursor = "hand2"
    ).pack(side = LEFT, padx = 5)

    Label(
        task_frame,
        text = f"{number}. Task: {task['task']}\n"
            f"Category: {task['category']}\n"
            f"Due Date: {task['date']}\n"
            f"Time: {task['time']}\n"
            f"Priority: {task['priority']}",
        justify = "left",
        anchor = "w",
        bg = box_colour,
        fg = text_colour
    ).pack(side = LEFT, padx = 5)

#This function use to clear the old task cards & create new one
def refresh_tasks():
    for widget in task_container.winfo_children():
        widget.destroy()

    #Display every task again
    for i, task in enumerate(tasks):
        display_task(task, i + 1)

    #Update the layout
    task_container.update_idletasks()

    #Update the scrolling area
    canvas.configure(scrollregion = canvas.bbox("all"))

def search_tasks():

    #Get what user search, convert it to lowercase, and remove spaces at the beginning & end
    keyword = search_entry.get().lower().strip()

    #If the search box empty, show all tasks
    if keyword == "":
        refresh_tasks()
        return
    
    for widget in task_container.winfo_children():
            widget.destroy()

    #Start numbering the result from 1
    result_number = 1

    #Check if it matches, when the condition is True the task is displayed
    for task in tasks:
        if (keyword in task["task"].lower() or
            keyword in task["category"].lower()):

            #Only tasks that match with what the user search will appear
            display_task(task, result_number)
            result_number += 1
    
    task_container.update_idletasks()
    canvas.configure(scrollregion = canvas.bbox("all"))

search_button = Button(search_frame,
    text = "Search",
    command = search_tasks,
    cursor = "hand2"
)
search_button.pack(side = LEFT, padx = 5)
        
show_button = Button(search_frame,
    text = "Show All",
    command = refresh_tasks,
    cursor = "hand2"
)
show_button.pack(side = LEFT, padx = 5)

#Create canvas so we can scroll in list_frame

canvas = Canvas(list_frame)
scrollbar = Scrollbar(list_frame, orient = "vertical", command = canvas.yview)

task_container = Frame(canvas)

#Mouse Tasks

def mouse_scroll(event):
    first, last = canvas.yview()

    if event.delta > 0 and first <= 0:
        return

    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", mouse_scroll))
canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

task_container.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion = canvas.bbox("all")
    )
)

canvas_window = canvas.create_window((0, 0), window = task_container, anchor = "nw")

#This function runs when the canvas change size
def resize_canvas(event):
    canvas.itemconfig(
        canvas_window,
        width = canvas.winfo_width()
    )

#When the widget change size, it'll call resize_canvas
canvas.bind("<Configure>", resize_canvas)

#Start scroll at the very top and connect the scrollbar with the canvas
canvas.yview_moveto(0)
canvas.configure(yscrollcommand = scrollbar.set)

canvas.pack(side = LEFT, fill = BOTH, expand = True)
scrollbar.pack(side = RIGHT, fill = Y)

#After 100 ms, the window has finished loading so the resigze_canvas get the correct width
root.after(100, lambda: resize_canvas(None))

#Functions

#The delete buttons are hidden
delete_mode = False

#This function use to delete one task from the task list
def delete_task(task_data):
    if task_data in tasks:
        tasks.remove(task_data)

    save_task()
    refresh_tasks()
    update_percentage()
    update_upcoming()
    update_overdue()
    update_buddy()

#This function use to turn Delete Mode on/off
def toggle_delete():
    global delete_mode

    #Like a switch, if before = False then after = True
    delete_mode = not delete_mode

    for task in tasks:

        #If the delete mode is on, show the delete button. If not then hide it
        if delete_mode:
            task["delete_button"].pack(side = RIGHT, padx = 5)
        else:
            task["delete_button"].pack_forget()

#This function use to return back to the welcome page
def go_back():

    update_streak()
    update_percentage()

    main_frame.pack_forget()
    welcome_frame.pack(fill = BOTH, expand = True)

#This function use to load every saved data before
def load_task():
    global tasks, streak, last_streak_date

    try:
        wb = load_workbook("tasks.xlsx")

        #Check if the workbook contain a sheet name Tasks, if yes then use that sheet
        if "Tasks" in wb.sheetnames:
            ws = wb["Tasks"]
        
        #If not, use whichever sheet that's currently active
        else:
            ws = wb.active        

        #We need this so we won't get duplicate tasks 
        tasks.clear()

        for widget in task_container.winfo_children():
            widget.destroy()

        #Read the worksheet row by row, start from row 2 because row 1 contain headings & return only values
        for row in ws.iter_rows(min_row = 2, values_only = True):

            #Create task dictionary
            task = {
                "task": row[0],
                "category": row[1],
                "date": row[2],
                "time": row[3],
                "priority": row[4],
                "done": row[5],

                #If len(row) > 6 is False, the Python doesn't try to read row[6]. This avoids an IndexError
                "completed_date": row[6] if len(row) > 6 and row[6] else ""
            }

            #Add the dictionary into the tasks list
            tasks.append(task)

        #Check if the workbook contain Stats sheet
        if "Stats" in wb.sheetnames:

            stats = wb["Stats"]

            #Load streak
            streak = stats["B1"].value or 0

            #Load last streak date
            saved_date = stats["B2"].value

            if saved_date:
                last_streak_date = datetime.strptime(
                    saved_date,
                    "%m-%d-%Y"
                ).date()
            else:
                #No streak exists yet
                last_streak_date = None

        update_streak()

        refresh_tasks()
        update_percentage()
        update_upcoming()
        update_overdue()
        update_buddy()
        update_daily_quote()

    #If there is no tasks.xlsx, it will print no saved file found
    except FileNotFoundError:
        print("No saved file found.")

#This function use to save all the data into the Excel file
def save_task():

    #Create new workbook
    wb = Workbook()

    #Get the first worksheet and rename it to Tasks
    ws = wb.active
    ws.title = "Tasks"      

    #Create the column headings. append() add one row
    ws.append([
        "Task",
        "Category",
        "Date",
        "Time",
        "Priority",
        "Done",
        "Completed Date"
    ])

    #Save every task
    for task in tasks:
        ws.append([
            task["task"],
            task["category"],
            task["date"],
            task["time"],
            task["priority"],
            task["done"],
            task["completed_date"]
        ])

    #Create another worksheet called Stats
    stats = wb.create_sheet("Stats")

    #Cell A1 become Streak and save the streak value to B1
    stats["A1"] = "Streak"
    stats["B1"] = streak

    #Write another label which is last date
    stats["A2"] = "Last Date"

    #Save the last streak date
    if last_streak_date:
        stats["B2"] = last_streak_date.strftime("%m-%d-%Y")
    else:
        stats["B2"] = ""

    #Save the workbook
    wb.save("tasks.xlsx")
    print("Saved successfully!")

#Popups

#No popup window at first
add_window = None

#This function run when the user click the plus button
def open_add_task():
    global add_window

    #Check if popup already created or closed, if both true the popup exist
    if add_window is not None and add_window.winfo_exists():

        #Move the popup to the front instead of creating another one
        add_window.lift()

        #Get the keyboard focus to that window
        add_window.focus_force()

        #Stop the function so the user won't accidentally open many Add Task windows
        return

    #Create a new window that attached to the main 
    add_window = Toplevel(root)
    add_window.title("Add Task")
    add_window.geometry("400x600")


    def close_add_window():
        global add_window

        add_window.destroy()

        #Tell the program that the popup no longer exist
        add_window = None
    
    #When the user click X, it will run close_add_window, destroy the window then make the add_window = None
    add_window.protocol("WM_DELETE_WINDOW", close_add_window)

    Label(add_window,
        text = "Task Name: ",
        font = ("Helvetica", 14)
    ).pack(pady = 10)
    
    task_entry = Entry(add_window)
    task_entry.pack()

    Label(add_window,
        text = "Category: ",
        font = ("Helvetica", 14)
    ).pack(pady = 10)

    category_entry = Entry(add_window)
    category_entry.pack()

    Label(add_window,
        text = "Due Date (MM-DD-YYYY): ",
        font = ("Helvetica", 14)
    ).pack(pady = 10)

    date_entry = Entry(add_window)
    date_entry.pack()

    Label(add_window,
        text = "Time (HH:MM): ",
        font = ("Helvetica", 14)
    ).pack(pady = 10)

    time_entry = Entry(add_window)
    time_entry.pack()

    Label(add_window,
        text = "Priority: ",
        font = ("Helvetica", 14)
    ).pack(pady = 10)

    priority_var = StringVar(value = "Medium")

    #To choose the priority level
    Radiobutton(
        add_window,
        text = "High",
        variable = priority_var,
        value = "High",
        cursor = "hand2"
    ).pack()

    Radiobutton(
        add_window,
        text = "Medium",
        variable = priority_var,
        value = "Medium",
        cursor = "hand2"
    ).pack()

    Radiobutton(
        add_window,
        text = "Low",
        variable = priority_var,
        value = "Low",
        cursor = "hand2"
    ).pack()
    
    #This function run when the Add Task button clicked
    def add_task():

        #Read the user input
        task = task_entry.get()
        category = category_entry.get()
        date = date_entry.get()
        time = time_entry.get()
        priority = priority_var.get()

        #Validate the date, is the date format correct or not
        try:
            datetime.strptime(date, "%m-%d-%Y")
        except ValueError:
            messagebox.showerror(
                "Invalid Date",
                "PLease enter date as MM-DD-YYYY"
            )
            return
        
        #Validate the time, is the time format correct or not
        try:
            datetime.strptime(time, "%H:%M")
        except ValueError:
            messagebox.showerror(
                "Invalid Time",
                "Please enter time as HH:MM"
            )
            return
        
        #Combine date and time into datetime
        task_datetime = datetime.strptime(
            f"{date} {time}",
            "%m-%d-%Y %H:%M"
        )

        #If user input date that already passed, it will show messagebox
        if task_datetime <= datetime.now():
            messagebox.showerror(
                "Expired Task",
                "This date and time have already passed."
            )
            return
        
        #Add new dictionary
        tasks.append({
                "task": task,
                "category": category,
                "date": date,
                "time": time,
                "priority": priority,
                "done": False,
                "completed_date": ""
            })

        save_task()
        refresh_tasks()
        update_percentage()
        update_upcoming()
        update_overdue()
        update_buddy()

        close_add_window()

    Button(
        add_window,
        text = "Add Task",
        command = add_task,
        cursor = "hand2"
    ).pack(pady=10)

#Buttons

button_frame = Frame(main_frame)
button_frame.pack(fill = X, pady = 10)

delete_button = Button(button_frame, 
    text = "Delete",
    command = toggle_delete,
    cursor = "hand2"
)
delete_button.pack(side = LEFT, padx = 10) 

back_button = Button(button_frame,
    text = "Back",
    command = go_back,
    cursor = "hand2"
)
back_button.pack(side = LEFT, padx = 10)

plus_button = Button(
    list_frame,
    image = plus_photo,
    command = open_add_task,
    bd = 0,                     
    highlightthickness = 0,     
    relief = "flat",
    cursor = "hand2"
)
plus_button.place(relx = 0.9, rely = 0.9, anchor = "se")

#Characters

buddy_frame = Frame(
    list_frame,
    bd = 0
)

buddy_frame.place(
    relx = 0.67,
    rely = 0.45,
    anchor = "center"
)

motivation = [
    "You got this! 💪",
    "Small steps every day 🌱",
    "Keep going! ⭐",
    "Don't forget to drink water!🥤"
]
motivation_job = None

#This function use to show random motivational messages
def show_motivation():
    global motivation_job

    #Coun the progress
    completed = sum(task["done"] for task in tasks)
    total = len(tasks)

    #Calculate how long the user has been inactive
    idle_seconds = (datetime.now() - last_activity).total_seconds()

    #If the user hasn't interact with the app for 60 secs, stop the function
    if idle_seconds > 60:
        return

    #Atleast one task completed
    if 0 < completed < total:
        
        #Pick random motivational messages from the motivation list
        speech_label.config(
            text = random.choice(motivation)
        )

        #Wait for 4 secs before run again
        motivation_job = root.after(
            4000,
            show_task_message
        )

speech_label = Label(
    buddy_frame,
    text = "Let's go! 🌱",
    font = ("Helvetica",12),
    padx = 15,
    pady = 10,
    wraplength = 180,
    relief = "ridge",
    bd = 2
)
speech_label.pack(pady = (0,10))

buddy_label = Label(
    buddy_frame,
    image = char_photo,
)
buddy_label.pack()

#This function use to update the buddy's img and messages
def update_buddy():
    global motivation_job

    #If the user completes another task before 4 secs, cancel the old timer
    if motivation_job is not None:
        root.after_cancel(motivation_job)
        motivation_job = None

    completed = sum(task["done"] for task in tasks)
    total = len(tasks)

    if dark_mode:
        speech_label.config(
            bg = "#3A3A3A",
            fg = "#FFFFFF"
        )
    else:
        speech_label.config(
            bg = "#FFF8D6",
            fg = "#000000"
        )

    if total == 0:
        buddy_label.config(image = char_photo)
        speech_label.config(
            text = "Let's plan today's tasks! 🌱"
        )

    elif completed == 0:
        buddy_label.config(image = bello_photo)
        speech_label.config(
            text = "Ready to start? 💪"
        )

    elif completed < total:
        buddy_label.config(image = happy_photo)

        show_task_message()

    else:
        buddy_label.config(image = happy_photo)
        speech_label.config(
            text = "YOU DID IT!! 🎉"
        )

#This function use to make the buddy alternate between task reminder and motivational message, but only while the user is active
def show_task_message():
    global motivation_job

    completed = sum(task["done"] for task in tasks)
    total = len(tasks)

    idle_seconds = (datetime.now() - last_activity).total_seconds()

    #Check id the user is idle, if it's over 60 the function exit immediately
    if idle_seconds > 60:
        return

    #Atleast one task completed
    if 0 < completed < total:

        speech_label.config(
            text = f"Only {total-completed} task(s) left! ⭐"
        )

        #After 4 secs, Tkinter call show_motivation. So the buddy alternate
        motivation_job = root.after(
            4000,
            show_motivation
        )

#Store the time of the user's most recent interaction
last_activity = datetime.now()

#This function run when the user click a key or the mouse
def user_activity(event = None):
    global last_activity

    #Calculate whether the user was idle
    was_idle = (
        datetime.now() - last_activity
    ).total_seconds() > 60

    #Update activity time so the app know the user is active again
    last_activity = datetime.now()

    #If they just returned from being idle, it will call update_buddy
    if was_idle:
        update_buddy()

#Whenever a key pressed and mouse clicked, it will call user_activity
root.bind_all("<Key>", user_activity)
root.bind_all("<Button-1>", user_activity)

#This function run every second to check whether the user has been inactive for more than 60 secs
def check_idle():
    global motivation_job

    #Calculate idle time
    idle_seconds = (
        datetime.now() - last_activity
    ).total_seconds()

    #If idle for more than 60 secs, the buddy goes to sleep 
    if idle_seconds > 60:
        
        #If the user is idle, stop the scheduled callback
        if motivation_job is not None:
            root.after_cancel(motivation_job)
            motivation_job = None

        buddy_label.config(image = sleep_photo)
        speech_label.config(
            text="Zzz... 😴"
        )

    #Continously check whether the user is idle
    root.after(1000, check_idle)

floating = True

#This function use to create simple up and down animation
def float_buddy():
    global floating

    #If floating = true, the buddy placed slightly lower. If = false, the buddy placed slightly higher
    if floating:
        buddy_frame.place_configure(rely = 0.45)
    else:
        buddy_frame.place_configure(rely = 0.44)

    floating = not floating

    #Switch direction. True become False, False become True
    root.after(500, float_buddy)

#Begin the animation
float_buddy()

#Count completed tasks
completed = 0

for task in tasks:
    if task["done"]:
        completed += 1

update_clock()

load_task()       
update_streak()

check_idle()
check_reminders()

#Start Tkinter's event loop
root.mainloop()